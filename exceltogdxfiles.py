# Caveat: Only runs on Windows with Excel installed due to gdxxrw usage!

import openpyxl
import os


def del_if_exists(sheet_name, wb):
    if sheet_name in wb:
        del wb[sheet_name]


def write_gdx_for_project_data(l, wb):
    worksheet_name = 'Projekt ' + str(l + 1)
    ws = wb[worksheet_name]

    wb2 = openpyxl.load_workbook('RCPSPPSinputTemplate.xlsx')

    del_if_exists('Demo', wb2)
    ws2 = wb2.create_sheet('Demo')

    for row in ws.rows:
        for cell in row:
            ws2[cell.coordinate].value = cell.value

    del_if_exists('Sheet', wb2)
    out_filename = 'RCPSPPSinputProject' + str(l + 1) + '.xlsx'
    wb2.save(out_filename)
    out_gdx_filename = 'Projekt' + str(l + 1) + '.gdx'
    os.system('gdxxrw i=' + out_filename + ' o=' + out_gdx_filename + ' @inputformat.txt')
    os.remove(out_filename)


def main():
    NUM_PROJECTS = 3

    wb = openpyxl.load_workbook('Input.xlsx')

    input_format_str = '''set=j       rng=Sets!A5 Rdim=1
    set=t       rng=Sets!A20 Rdim=1
    set=r       rng=Sets!C20 Rdim=1
    set=n       rng=Sets!E20 Rdim=1
    set=e       rng=Sets!C5 Rdim=1
    set=V       rng=Parameters!R5 Rdim=1 values=yn
    
    set=a       rng=Tables!A5:C15 Rdim=1 Cdim=1 values=yn
    set=B       rng=Tables!E5:O15 Rdim=1 Cdim=1 values=yn
    set=P       rng=Tables!E18:P28 Rdim=1 Cdim=1 values=yn
    set=W       rng=Tables!T5:AD7 Rdim=1 Cdim=1 values=yn
    
    par=kr      rng=Tables!BV5:BW16 Rdim=1 Cdim=1
    par=kn      rng=Tables!BY5:BZ16 Rdim=1 Cdim=1
    
    par=c       rng=Parameters!W5 Rdim=0
    par=M       rng=Parameters!D5 Rdim=0
    par=s       rng=Parameters!U5 Rdim=0
    par=d       rng=Parameters!I5 Rdim=1'''

    input_format_str2 = '''
    par=Kr_m  rng=Globals!B3 Rdim=1
    par=Kn_m  rng=Globals!E3 Rdim=1
    '''

    with open('inputformat.txt', 'w') as fp:
        fp.write(input_format_str)
    with open('inputformat2.txt', 'w') as fp:
        fp.write(input_format_str2)

    for l in range(NUM_PROJECTS): write_gdx_for_project_data(l, wb)
    os.system('gdxxrw i=Input.xlsx o=Capacities.gdx @inputformat2.txt')


if __name__ == '__main__': main()
