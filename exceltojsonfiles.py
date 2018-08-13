import openpyxl
import json


def column_as_list(ws, column_name, start_row_index, end_row_index):
    return [cell.value for cell in ws[column_name][start_row_index - 1:end_row_index]]


def indicator_column_to_sublist(ws, column_name, start_row_index, end_row_index):
    return [row_ix - start_row_index + 1 for row_ix in range(start_row_index, end_row_index + 1) if
            ws[column_name][row_ix - 1].value == 'yes']


def rect_map(ws, f, top_left_indices, bottom_right_indices):
    tl_x, tl_y = top_left_indices
    br_x, br_y = bottom_right_indices
    return [[f(cell) for cell in row] for row in
            ws.iter_rows(min_row=tl_x, max_row=br_x, min_col=tl_y, max_col=br_y)]


def rect_as_list_of_rows(ws, top_left_indices, bottom_right_indices):
    return rect_map(ws, lambda cell: cell.value, top_left_indices, bottom_right_indices)


def rect_as_list_of_boolean_rows(ws, top_left_indices, bottom_right_indices):
    return rect_map(ws, lambda cell: cell.value == 'yes', top_left_indices, bottom_right_indices)


def write_as_json(dict, out_filename):
    with open(out_filename, 'w') as fp:
        fp.write(json.dumps(dict, sort_keys=True, indent=4))


def convert_excel_to_project_jsons(input_filename = 'Input.xlsx', num_projects = 3):
    wb = openpyxl.load_workbook(input_filename)
    projects = []

    for i in range(num_projects):
        worksheet_name = 'Projekt ' + str(i + 1)
        ws = wb[worksheet_name]

        projects.append({
            'jobs': column_as_list(ws, 'A', 6, 15),
            'durations': column_as_list(ws, 'B', 6, 15),
            'demands': [column_as_list(ws, 'C', 6, 15)],
            'demands_nonrenewable': [column_as_list(ws, 'D', 6, 15)],
            'capacities': [wb['Globals']['C3'].value, wb['Globals']['F3'].value],
            'mandatory_activities': indicator_column_to_sublist(ws, 'E', 6, 15),
            'job_in_decision': rect_as_list_of_boolean_rows(ws, (6, 6), (15, 6)),
            'job_activating_decision': rect_as_list_of_boolean_rows(ws, (6, 7), (15, 7)),
            'job_causing_job': rect_as_list_of_boolean_rows(ws, (6, 8), (15, 17)),
            'precedence': rect_as_list_of_boolean_rows(ws, (6, 18), (15, 27)),
            'deadline': ws['B18'].value,
            'delaycost': ws['B19'].value,
            'kappa': [wb['Globals']['L3'].value],
            'zmax': [wb['Globals']['I3'].value],
            'qlevel_requirement': rect_as_list_of_rows(ws, (23, 3), (24, 5)),
            'revenues': rect_as_list_of_rows(ws, (25, 3), (27, 5)),
            'revenue_periods': column_as_list(ws, 'B', 25, 27),
            'base_qualities': column_as_list(ws, 'F', 23, 24),
            'costs': column_as_list(ws, 'AD', 6, 15),
            'quality_improvements': rect_as_list_of_rows(ws, (6, 28), (15, 29))
        })

    return projects


def main():
    projects = convert_excel_to_project_jsons('Input.xlsx')
    for i, project in enumerate(projects):
        write_as_json(project, 'Projekt' + str(i + 1) + '.json')


if __name__ == '__main__': main()
