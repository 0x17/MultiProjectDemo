import openpyxl

import psplibtojson

NUM_PROJECTS = 3


def decision_indices(pobj):
    return list(pobj['jobsInDecision'].keys())


def optional_jobs(obj, l):
    pobj = obj['projects'][l]
    jobs_in_any_decision = []
    for d in decision_indices(pobj):
        jobs_in_any_decision += pobj['jobsInDecision'][d]
    jobs_caused_by_any_job = [v for k, v in pobj['jobCausingJob'].items()]
    return list(set(jobs_in_any_decision).union(set(jobs_caused_by_any_job)))


def mandatory_jobs(obj, l):
    return set(range(obj['projects'][l]['numJobs'])).difference(set(optional_jobs(obj, l)))


def fill_column_cells(ws, column_name, start_row_index, end_row_index, values):
    ctr = 0
    for cell in ws[column_name][start_row_index - 1:end_row_index]:
        cell.value = values[ctr]
        ctr += 1


def fill_rect(ws, top_left_indices, mx):
    for rix in range(len(mx)):
        for cix in range(len(mx[0])):
            ws.cell(row=rix+top_left_indices[0], column=cix+top_left_indices[1], value=mx[rix][cix])


def pair_in_dict(pair, dict):
    return pair[0] in dict and dict[pair[0]] == pair[1]


def fill_excel_with_obj(obj, template_fn, out_fn):
    wb = openpyxl.load_workbook(template_fn)
    for l in range(NUM_PROJECTS):
        ws = wb['Projekt ' + str(l + 1)]
        pobj = obj['projects'][l]
        njobs = pobj['numJobs']

        fill_column_cells(ws, 'B', 6, 15, pobj['durations'])
        fill_column_cells(ws, 'C', 6, 15, [dpair[0] for dpair in pobj['demands']])
        fill_column_cells(ws, 'D', 6, 15, [dpair[1] for dpair in pobj['demands']])
        fill_column_cells(ws, 'E', 6, 15, ['yes' if j in mandatory_jobs(obj, l) else '' for j in range(njobs)])

        fill_rect(ws, (6,6), [['yes' if j in pobj['jobsInDecision'][e] else '' for e in decision_indices(pobj) ] for j in range(njobs) ])
        fill_rect(ws, (6,7), [['yes' if pair_in_dict((str(j), e), pobj['jobCausingDecision']) else '' for e in decision_indices(pobj)] for j in range(njobs)])
        fill_rect(ws, (6,8), [['yes' if pair_in_dict((str(i), j), pobj['jobCausingJob']) else '' for j in range(njobs)] for i in range(njobs)])
        fill_rect(ws, (6,18), [['yes' if j in pobj['successors'][i] else '' for j in range(njobs)] for i in range(njobs)])

        ws['B18'].value = pobj['deadline']
        ws['B19'].value = pobj['delaycost']

    wb['Globals']['C3'].value = obj['Kr'][0]
    wb['Globals']['F3'].value = obj['Kn'][0]

    wb.save(out_fn)


def main():
    obj = psplibtojson.parse_json_from_psplib('Instanzen_Begehung/Modellendogen0003.DAT')
    fill_excel_with_obj(obj, 'InputTemplate.xlsx', 'Input.xlsx')


if __name__ == '__main__': main()
